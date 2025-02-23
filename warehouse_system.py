try:
    import tkinter as tk
    from tkinter import ttk, messagebox, filedialog
except ImportError as e:
    print("错误：无法导入tkinter模块。请确保Python正确安装并包含tkinter。")
    print("详细错误信息:", str(e))
    input("按回车键退出...")
    exit(1)

try:
    import sqlite3
    from datetime import datetime
    import pandas as pd  # 用于Excel导出
    import os  # 用于文件操作
except ImportError as e:
    print("错误：无法导入必要的模块。")
    print("详细错误信息:", str(e))
    print("请安装所需模块：")
    print("pip install pandas openpyxl")
    input("按回车键退出...")
    exit(1)

try:
    from tkcalendar import DateEntry
except ImportError as e:
    print("错误：无法导入tkcalendar模块。")
    print("详细错误信息:", str(e))
    print("请安装所需模块：")
    print("pip install tkcalendar")
    input("按回车键退出...")
    exit(1)

class WarehouseSystem:
    def __init__(self, root):
        self.root = root
        self.root.title("仓库管理系统")
        
        # 设置窗口大小和位置
        window_width = 800
        window_height = 600
        
        # 获取屏幕尺寸
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        # 计算窗口居中的坐标
        x = (screen_width - window_width) // 2
        y = (screen_height - window_height) // 2
        
        # 设置窗口大小和位置
        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        
        # 设置窗口最小尺寸
        self.root.minsize(800, 600)
        
        # 确保窗口在最前面显示
        self.root.lift()
        self.root.focus_force()
        
        # 创建标题框架
        title_frame = ttk.Frame(self.root)
        title_frame.pack(pady=10)
        
        # 添加logo
        try:
            # 使用 tkinter 的 PhotoImage 直接加载 png
            logo_photo = tk.PhotoImage(file='logo.png')  # 需要使用 PNG 格式的图片
            # 调整大小
            logo_photo = logo_photo.subsample(logo_photo.width()//30, logo_photo.height()//30)
            # 保存引用，防止被垃圾回收
            self.logo_photo = logo_photo
            logo_label = ttk.Label(title_frame, image=logo_photo, cursor='hand2')
            logo_label.pack(side=tk.LEFT, padx=5)
            logo_label.bind('<Button-1>', lambda e: self.show_about_info())
            
            # 设置窗口图标
            self.root.iconbitmap('logo.ico')
        except:
            # 如果加载失败，显示一个空白标签
            ttk.Label(title_frame, width=4).pack(side=tk.LEFT, padx=5)
        
        # 创建标题标签并绑定点击事件
        self.title_label = tk.Label(title_frame, text="仓库管理系统", 
                                   font=('黑体', 16, 'bold'),
                                   cursor='hand2')  # 设置鼠标悬停时的光标样式为手型
        self.title_label.pack(side=tk.LEFT, padx=5)
        self.title_label.bind('<Button-1>', lambda e: self.show_about_info())
        
        # 检查是否首次运行
        self.check_first_run()
        
        # 添加编辑模式标志
        self.editing_mode = False
        
        # 创建数据库连接
        self.create_database()
        
        # 创建主界面
        self.create_gui()
        
    def create_database(self):
        conn = sqlite3.connect('warehouse.db')
        cursor = conn.cursor()
        
        # 创建商品表
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            quantity INTEGER NOT NULL,
            price REAL,
            category TEXT
        )
        ''')
        
        # 修改交易记录表，添加经办人字段
        cursor.execute('''
        CREATE TABLE IF NOT EXISTS transactions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            product_id INTEGER,
            type TEXT NOT NULL,
            quantity INTEGER NOT NULL,
            operator TEXT NOT NULL,
            date TEXT NOT NULL,
            FOREIGN KEY (product_id) REFERENCES products (id)
        )
        ''')
        
        conn.commit()
        conn.close()
        
    def create_gui(self):
        # 创建菜单栏
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        
        # 添加"关于"菜单
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="帮助", menu=help_menu)
        help_menu.add_command(label="关于", command=self.show_about_info)
        
        # 创建选项卡（放在标题标签下方）
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(expand=True, fill='both')
        
        # 商品管理选项卡
        self.products_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.products_frame, text='     商品管理     ')  # 增加空格使标签更宽
        self.create_products_tab()
        
        # 出入库管理选项卡
        self.transactions_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.transactions_frame, text='     出入库管理     ')  # 增加空格使标签更宽
        self.create_transactions_tab()
        
    def create_products_tab(self):
        # 创建输入框框架
        input_frame = ttk.LabelFrame(self.products_frame, text="添加商品")
        input_frame.pack(fill="x", padx=5, pady=5)
        
        # 设置网格列的权重，使内容居中
        input_frame.grid_columnconfigure(0, weight=1)
        input_frame.grid_columnconfigure(3, weight=1)
        
        # 商品信息输入
        ttk.Label(input_frame, text="商品ID:", font=('微软雅黑', 10)).grid(row=0, column=0, padx=(10, 0), pady=8, sticky='e')
        self.id_entry = ttk.Entry(input_frame, justify='center', font=('微软雅黑', 10))
        self.id_entry.grid(row=0, column=1, padx=(5, 25), pady=8)
        self.id_entry.bind('<KeyRelease>', self.update_product_name_in_products)
        self.id_entry.bind('<Return>', self.handle_enter_key)
        
        ttk.Label(input_frame, text="商品名称:", font=('微软雅黑', 10)).grid(row=0, column=2, padx=(10, 0), pady=8, sticky='e')
        self.name_entry = ttk.Entry(input_frame, justify='center', font=('微软雅黑', 10))
        self.name_entry.grid(row=0, column=3, padx=(5, 10), pady=8)
        self.name_entry.bind('<KeyRelease>', self.update_product_id_in_products)
        self.name_entry.bind('<Return>', self.handle_enter_key)
        
        ttk.Label(input_frame, text="价格:", font=('微软雅黑', 10)).grid(row=1, column=0, padx=(10, 0), pady=8, sticky='e')
        self.price_entry = ttk.Entry(input_frame, justify='center', font=('微软雅黑', 10))
        self.price_entry.grid(row=1, column=1, padx=(5, 25), pady=8)
        
        ttk.Label(input_frame, text="类别:", font=('微软雅黑', 10)).grid(row=1, column=2, padx=(10, 0), pady=8, sticky='e')
        self.category_entry = ttk.Entry(input_frame, justify='center', font=('微软雅黑', 10))
        self.category_entry.grid(row=1, column=3, padx=(5, 10), pady=8)
        
        # 添加按钮框架
        button_frame = ttk.Frame(input_frame)
        button_frame.grid(row=2, column=0, columnspan=4, pady=12)
        
        style = ttk.Style()
        style.configure('Action.TButton', font=('微软雅黑', 10))
        
        ttk.Button(button_frame, text="添加商品", style='Action.TButton', command=self.add_product).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="编辑商品", style='Action.TButton', command=self.edit_product).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="删除商品", style='Action.TButton', command=self.delete_product).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="导出Excel", style='Action.TButton', command=lambda: self.export_to_excel("products")).pack(side=tk.LEFT, padx=10)
        
        # 创建表格
        self.tree = ttk.Treeview(self.products_frame, columns=("序列", "ID", "名称", "数量", "价格", "类别"), show="headings")
        self.tree.heading("序列", text="序列")
        self.tree.heading("ID", text="商品ID")
        self.tree.heading("名称", text="商品名称")
        self.tree.heading("数量", text="数量")
        self.tree.heading("价格", text="价格")
        self.tree.heading("类别", text="类别")
        
        # 设置列宽
        self.tree.column("序列", width=15)
        self.tree.column("ID", width=40)
        self.tree.column("名称", width=150)
        self.tree.column("数量", width=80)
        self.tree.column("价格", width=80)
        self.tree.column("类别", width=100)
        
        self.tree.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 设置标签样式 - 使用更深的颜色
        self.tree.tag_configure('oddrow', background='#E8E8E8')  # 更深的灰色
        self.tree.tag_configure('evenrow', background='#F8F8F8')  # 较浅的灰色
        
        # 绑定选择事件
        self.tree.bind('<<TreeviewSelect>>', self.item_selected)
        
        # 添加双击事件绑定
        self.tree.bind('<Double-1>', self.on_product_double_click)
        
        # 刷新商品列表
        self.refresh_products()
        
    def create_transactions_tab(self):
        # 创建输入框框架
        input_frame = ttk.LabelFrame(self.transactions_frame, text="出入库操作")
        input_frame.pack(fill="x", padx=5, pady=5)
        
        # 设置网格列的权重，使内容居中
        input_frame.grid_columnconfigure(0, weight=1)
        input_frame.grid_columnconfigure(3, weight=1)
        
        # 第一行：商品ID和名称
        ttk.Label(input_frame, text="商品ID:", font=('微软雅黑', 10)).grid(row=0, column=0, padx=(10, 0), pady=8, sticky='e')
        self.trans_product_id = ttk.Combobox(input_frame, width=12, justify='center', font=('微软雅黑', 10))
        self.trans_product_id.grid(row=0, column=1, padx=(5, 25), pady=8)
        
        ttk.Label(input_frame, text="商品名称:", font=('微软雅黑', 10)).grid(row=0, column=2, padx=(10, 0), pady=8, sticky='e')
        self.trans_product_name = ttk.Combobox(input_frame, width=22, justify='center', font=('微软雅黑', 10))
        self.trans_product_name.grid(row=0, column=3, padx=(5, 10), pady=8)
        
        # 初始化下拉列表的值
        try:
            conn = sqlite3.connect('warehouse.db')
            cursor = conn.cursor()
            
            # 初始化商品ID列表
            cursor.execute('SELECT id FROM products ORDER BY id')
            ids = [str(row[0]) for row in cursor.fetchall()]
            self.trans_product_id['values'] = ids
            
            # 初始化商品名称列表
            cursor.execute('SELECT name FROM products ORDER BY name')
            names = [row[0] for row in cursor.fetchall()]
            self.trans_product_name['values'] = names
            
            conn.close()
        except Exception as e:
            messagebox.showerror("错误", f"初始化商品列表失败：{str(e)}")
        
        # 绑定事件
        self.trans_product_id.bind('<FocusIn>', self.update_product_id_list)
        self.trans_product_name.bind('<FocusIn>', self.update_product_name_list)
        self.trans_product_id.bind('<<ComboboxSelected>>', self.on_product_id_selected)
        self.trans_product_name.bind('<<ComboboxSelected>>', self.on_product_name_selected)
        self.trans_product_id.bind('<KeyRelease>', self.update_product_name)
        self.trans_product_name.bind('<KeyRelease>', self.update_product_id)
        
        # 第二行：数量和经办人
        ttk.Label(input_frame, text="数量:", font=('微软雅黑', 10)).grid(row=1, column=0, padx=(10, 0), pady=8, sticky='e')
        self.trans_quantity = ttk.Entry(input_frame, width=12, justify='center', font=('微软雅黑', 10))
        self.trans_quantity.grid(row=1, column=1, padx=(5, 25), pady=8)
        
        ttk.Label(input_frame, text="经办人:", font=('微软雅黑', 10)).grid(row=1, column=2, padx=(10, 0), pady=8, sticky='e')
        self.trans_operator = ttk.Entry(input_frame, width=22, justify='center', font=('微软雅黑', 10))
        self.trans_operator.grid(row=1, column=3, padx=(5, 10), pady=8)
        
        # 第三行：日期
        ttk.Label(input_frame, text="日期:", font=('微软雅黑', 10)).grid(row=2, column=0, padx=(10, 0), pady=8, sticky='e')
        self.trans_date = ttk.Entry(input_frame, justify='center', font=('微软雅黑', 10))
        self.trans_date.grid(row=2, column=1, columnspan=2, padx=(5, 25), pady=8, sticky='ew')
        self.trans_date.insert(0, datetime.now().strftime('%Y-%m-%d'))
        
        # 修改按钮框架
        button_frame = ttk.Frame(input_frame)
        button_frame.grid(row=3, column=0, columnspan=4, pady=12)
        
        style = ttk.Style()
        style.configure('Action.TButton', font=('微软雅黑', 10))
        
        ttk.Button(button_frame, text="入库", style='Action.TButton', command=lambda: self.add_transaction("入库")).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="出库", style='Action.TButton', command=lambda: self.add_transaction("出库")).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="导出Excel", style='Action.TButton', command=lambda: self.export_to_excel("transactions")).pack(side=tk.LEFT, padx=10)
        
        # 创建表格
        self.trans_tree = ttk.Treeview(self.transactions_frame, 
                                     columns=("序列", "商品ID", "商品名称", "类型", "数量", "经办人", "日期"), 
                                     show="headings")
        self.trans_tree.heading("序列", text="序列")
        self.trans_tree.heading("商品ID", text="商品ID")
        self.trans_tree.heading("商品名称", text="商品名称")
        self.trans_tree.heading("类型", text="类型")
        self.trans_tree.heading("数量", text="数量")
        self.trans_tree.heading("经办人", text="经办人")
        self.trans_tree.heading("日期", text="日期")
        
        # 设置列宽
        self.trans_tree.column("序列", width=30)
        self.trans_tree.column("商品ID", width=50)
        self.trans_tree.column("商品名称", width=180)
        self.trans_tree.column("类型", width=70)
        self.trans_tree.column("数量", width=70)
        self.trans_tree.column("经办人", width=100)
        self.trans_tree.column("日期", width=100)
        
        self.trans_tree.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 设置标签样式 - 使用更深的颜色
        self.trans_tree.tag_configure('oddrow', background='#E8E8E8')  # 更深的灰色
        self.trans_tree.tag_configure('evenrow', background='#F8F8F8')  # 较浅的灰色
        
        # 在创建trans_tree后添加双击事件绑定
        self.trans_tree.bind('<Double-1>', self.on_tree_double_click)
        
        # 添加单击事件绑定
        self.trans_tree.bind('<ButtonRelease-1>', self.on_transaction_click)
        
        # 刷新交易记录
        self.refresh_transactions()
        
    def handle_enter_key(self, event):
        if self.editing_mode:
            self.edit_product()
        else:
            self.add_product()

    def add_product(self):
        try:
            name = self.name_entry.get().strip()
            product_id = self.id_entry.get().strip()
            
            # 验证商品ID是否为数字
            if not product_id.isdigit():
                messagebox.showerror("错误", "商品ID只能输入数字")
                return
            
            if not name:
                messagebox.showerror("错误", "请输入商品名称")
                return
            
            # 检查商品ID是否已存在
            conn = sqlite3.connect('warehouse.db')
            cursor = conn.cursor()
            
            cursor.execute('SELECT id FROM products WHERE id=?', (product_id,))
            existing_id = cursor.fetchone()
            
            cursor.execute('SELECT id FROM products WHERE name=?', (name,))
            existing_name = cursor.fetchone()
            
            if existing_id:
                messagebox.showerror("错误", "该商品ID已存在，请使用其他ID")
                conn.close()
                return
            
            if existing_name:
                messagebox.showerror("错误", "该商品名称已存在，请使用其他名称")
                conn.close()
                return
            
            # 新商品的数量默认为0
            quantity = 0
            
            try:
                price = float(self.price_entry.get()) if self.price_entry.get().strip() else None
            except ValueError:
                messagebox.showerror("错误", "价格必须是数字")
                return
            
            category = self.category_entry.get() if self.category_entry.get().strip() else None
            
            cursor.execute('''
            INSERT INTO products (id, name, quantity, price, category)
            VALUES (?, ?, ?, ?, ?)
            ''', (product_id, name, quantity, price, category))
            conn.commit()
            conn.close()
            
            self.refresh_products()
            self.clear_entries()
            messagebox.showinfo("成功", "商品添加成功！")
            
        except sqlite3.IntegrityError:
            messagebox.showerror("错误", "商品ID已存在，请使用其他ID")
        except ValueError:
            messagebox.showerror("错误", "输入格式错误，请检查输入内容")
        except Exception as e:
            messagebox.showerror("错误", f"添加商品时出错：{str(e)}")
            
    def delete_product(self):
        try:
            selected = self.tree.selection()
            if not selected:
                messagebox.showerror("错误", "请选择要删除的商品")
                return
                
            if not messagebox.askyesno("确认", "确定要删除选中的商品吗？"):
                return
                
            product_id = self.tree.item(selected[0])['values'][1]
            
            conn = sqlite3.connect('warehouse.db')
            cursor = conn.cursor()
            cursor.execute('DELETE FROM products WHERE id=?', (product_id,))
            conn.commit()
            conn.close()
            
            self.refresh_products()
            self.clear_entries()
            messagebox.showinfo("成功", "商品删除成功")
            
        except Exception as e:
            messagebox.showerror("错误", str(e))
            
    def add_transaction(self, trans_type):
        try:
            product_id = int(self.trans_product_id.get())
            quantity = int(self.trans_quantity.get())
            operator = self.trans_operator.get()
            trans_date = self.trans_date.get()
            
            # 只在出库时验证经办人
            if trans_type == "出库" and not operator.strip():
                messagebox.showerror("错误", "出库时请输入经办人")
                return
            
            # 如果是入库且没有填写经办人，设为空字符串
            if trans_type == "入库" and not operator.strip():
                operator = ""
            
            # 验证日期格式
            try:
                datetime.strptime(trans_date, '%Y-%m-%d')
            except ValueError:
                messagebox.showerror("错误", "日期格式错误，请使用格式：YYYY-MM-DD")
                return
            
            conn = sqlite3.connect('warehouse.db')
            cursor = conn.cursor()
            
            # 检查商品是否存在
            cursor.execute('SELECT quantity FROM products WHERE id=?', (product_id,))
            result = cursor.fetchone()
            
            if not result:
                messagebox.showerror("错误", "商品不存在")
                return
                
            current_quantity = result[0]
            
            # 出库时检查库存
            if trans_type == "出库" and quantity > current_quantity:
                messagebox.showerror("错误", "库存不足")
                return
                
            # 更新库存
            new_quantity = current_quantity + quantity if trans_type == "入库" else current_quantity - quantity
            cursor.execute('UPDATE products SET quantity=? WHERE id=?', (new_quantity, product_id))
            
            # 记录交易
            cursor.execute('''
            INSERT INTO transactions (product_id, type, quantity, operator, date)
            VALUES (?, ?, ?, ?, ?)
            ''', (product_id, trans_type, quantity, operator, trans_date))
            
            conn.commit()
            conn.close()
            
            self.refresh_products()
            self.refresh_transactions()
            self.clear_transaction_entries()
            messagebox.showinfo("成功", f"{trans_type}操作成功")
            
        except Exception as e:
            messagebox.showerror("错误", str(e))
            
    def refresh_products(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        conn = sqlite3.connect('warehouse.db')
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM products')
        for i, row in enumerate(cursor.fetchall(), 1):
            tag = 'evenrow' if i % 2 == 0 else 'oddrow'
            values = (i,) + row  # 添加序列号
            self.tree.insert('', 'end', values=values, tags=(tag,))
        conn.close()
        
    def refresh_transactions(self):
        for item in self.trans_tree.get_children():
            self.trans_tree.delete(item)
            
        conn = sqlite3.connect('warehouse.db')
        cursor = conn.cursor()
        cursor.execute('''
        SELECT t.id, t.product_id, p.name, t.type, t.quantity, t.operator, t.date 
        FROM transactions t 
        LEFT JOIN products p ON t.product_id = p.id 
        ORDER BY t.date DESC
        ''')
        for i, row in enumerate(cursor.fetchall(), 1):
            tag = 'evenrow' if i % 2 == 0 else 'oddrow'
            values = (i,) + row[1:]  # 用序列号替换ID
            self.trans_tree.insert('', 'end', values=values, tags=(tag,))
        conn.close()
        
    def item_selected(self, event):
        selected = self.tree.selection()
        if selected:
            # 进入编辑模式
            self.editing_mode = True
            
            item = self.tree.item(selected[0])
            values = item['values']
            self.id_entry.delete(0, tk.END)
            self.id_entry.insert(0, values[1])  # 商品ID
            self.name_entry.delete(0, tk.END)
            self.name_entry.insert(0, values[2])  # 商品名称
            self.price_entry.delete(0, tk.END)
            self.price_entry.insert(0, values[4])  # 价格
            self.category_entry.delete(0, tk.END)
            self.category_entry.insert(0, values[5])  # 类别

    def clear_entries(self):
        self.id_entry.delete(0, tk.END)
        self.name_entry.delete(0, tk.END)
        self.price_entry.delete(0, tk.END)
        self.category_entry.delete(0, tk.END)
        # 退出编辑模式
        self.editing_mode = False

    def update_product_name_in_products(self, event=None):
        """根据ID更新商品名称"""
        try:
            product_id = self.id_entry.get().strip()
            # 如果输入为空，清空商品名称并返回
            if not product_id:
                self.name_entry.delete(0, tk.END)
                return
            
            # 检查是否为数字
            if not product_id.isdigit():
                self.name_entry.delete(0, tk.END)
                messagebox.showwarning("提示", "商品ID只能输入数字")
                return
            
            conn = sqlite3.connect('warehouse.db')
            cursor = conn.cursor()
            cursor.execute('SELECT name FROM products WHERE id=?', (product_id,))
            result = cursor.fetchone()
            conn.close()
            
            self.name_entry.delete(0, tk.END)
            if result:
                self.name_entry.insert(0, result[0])
        except Exception as e:
            messagebox.showerror("错误", f"查询商品信息时出错：{str(e)}")

    def update_product_id_in_products(self, event=None):
        # 如果在编辑模式下，不执行自动填充
        if self.editing_mode:
            return
            
        try:
            product_name = self.name_entry.get()
            if product_name:
                conn = sqlite3.connect('warehouse.db')
                cursor = conn.cursor()
                cursor.execute('SELECT id FROM products WHERE name = ?', (product_name,))
                result = cursor.fetchone()
                conn.close()
                
                if result:
                    self.id_entry.delete(0, tk.END)
                    self.id_entry.insert(0, str(result[0]))
        except:
            pass

    def edit_product(self):
        try:
            selected = self.tree.selection()
            if not selected:
                messagebox.showerror("错误", "请选择要编辑的商品")
                return
            
            # 获取当前选中的商品ID和原始值
            item = self.tree.item(selected[0])
            values = item['values']
            product_id = values[1]
            original_name = values[2]
            original_price = values[4] if values[4] is not None else ""
            original_category = values[5] if values[5] is not None else ""
            
            # 获取输入框的当前值
            current_name = self.name_entry.get().strip()
            current_price = self.price_entry.get().strip()
            current_category = self.category_entry.get().strip()
            current_id = self.id_entry.get().strip()
            
            # 检查哪些字段被修改了
            updates = []
            update_values = []
            
            # 检查ID是否被修改
            if str(current_id) != str(product_id):
                # 检查新ID是否已存在
                conn = sqlite3.connect('warehouse.db')
                cursor = conn.cursor()
                cursor.execute('SELECT id FROM products WHERE id=?', (current_id,))
                if cursor.fetchone():
                    messagebox.showerror("错误", "该商品ID已存在")
                    conn.close()
                    return
                conn.close()
                
                if not current_id:
                    messagebox.showerror("错误", "商品ID不能为空")
                    return
                try:
                    int(current_id)
                except ValueError:
                    messagebox.showerror("错误", "商品ID必须为数字")
                    return
                updates.append("id=?")
                update_values.append(current_id)
            
            # 检查名称是否被修改
            if current_name != str(original_name):
                if not current_name:
                    messagebox.showerror("错误", "商品名称不能为空")
                    return
                # 检查新名称是否已存在（排除当前商品）
                conn = sqlite3.connect('warehouse.db')
                cursor = conn.cursor()
                cursor.execute('SELECT id FROM products WHERE name=? AND id!=?', (current_name, product_id))
                if cursor.fetchone():
                    messagebox.showerror("错误", "该商品名称已存在")
                    conn.close()
                    return
                conn.close()
                updates.append("name=?")
                update_values.append(current_name)
            
            # 只有当价格字段被修改时才验证
            if current_price != str(original_price):
                if current_price:  # 只有当用户输入了新价格时才验证
                    try:
                        price = float(current_price)
                        if price < 0:
                            messagebox.showerror("错误", "价格不能为负数")
                            return
                        updates.append("price=?")
                        update_values.append(price)
                    except ValueError:
                        messagebox.showerror("错误", "价格必须是数字")
                        return
                else:  # 如果用户清空了价格
                    updates.append("price=?")
                    update_values.append(None)
            
            # 检查类别是否被修改
            if current_category != str(original_category):
                updates.append("category=?")
                update_values.append(current_category if current_category else None)
            
            # 如果没有任何修改，直接返回
            if not updates:
                messagebox.showinfo("提示", "没有检测到任何修改")
                return
            
            # 确认是否更新
            if not messagebox.askyesno("确认", "确定要更新商品信息吗？"):
                return
            
            conn = sqlite3.connect('warehouse.db')
            cursor = conn.cursor()
            
            try:
                # 开始事务
                conn.execute("BEGIN TRANSACTION")
                
                # 构建更新SQL语句
                update_sql = "UPDATE products SET " + ", ".join(updates) + " WHERE id=?"
                update_values.append(product_id)
                
                # 执行商品表更新
                cursor.execute(update_sql, tuple(update_values))
                
                # 如果ID被修改了，更新交易记录表中的product_id
                if str(current_id) != str(product_id):
                    cursor.execute('UPDATE transactions SET product_id=? WHERE product_id=?', 
                                 (current_id, product_id))
                
                # 提交事务
                conn.commit()
                
                # 刷新显示
                self.refresh_products()
                self.refresh_transactions()
                self.clear_entries()
                messagebox.showinfo("成功", "商品信息已更新")
                
            except Exception as e:
                # 如果出现错误，回滚事务
                conn.rollback()
                raise e
            finally:
                conn.close()
            
        except Exception as e:
            messagebox.showerror("错误", str(e))

    def on_product_double_click(self, event):
        # 获取双击的项目和列
        item = self.tree.selection()[0]
        column = self.tree.identify_column(event.x)
        values = self.tree.item(item)['values']
        
        # 如果双击的是商品ID或商品名称列
        if column == '#2' or column == '#3':  # ID或名称列
            self.show_product_outbound_stats(values[1], values[2])  # 传递商品ID和名称

    def show_product_outbound_stats(self, product_id, product_name):
        # 创建新窗口
        stats_window = tk.Toplevel(self.root)
        stats_window.title(f"商品 {product_name} 的出库统计")
        
        # 设置窗口大小和位置
        stats_window.geometry("600x400")
        x = self.root.winfo_x() + (self.root.winfo_width() - 600) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 400) // 2
        stats_window.geometry(f"+{x}+{y}")
        
        # 创建主框架
        main_frame = ttk.Frame(stats_window)
        main_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 创建表格框架
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 创建统计表格
        stats_tree = ttk.Treeview(tree_frame, 
                                 columns=("经办人", "总出库数量", "出库次数"),
                                 show="headings")
        
        # 设置表头
        stats_tree.heading("经办人", text="经办人")
        stats_tree.heading("总出库数量", text="总出库数量")
        stats_tree.heading("出库次数", text="出库次数")
        
        # 设置列宽
        stats_tree.column("经办人", width=100)
        stats_tree.column("总出库数量", width=80)
        stats_tree.column("出库次数", width=80)
        
        # 添加垂直滚动条
        y_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=stats_tree.yview)
        stats_tree.configure(yscrollcommand=y_scrollbar.set)
        
        # 布局表格和滚动条
        stats_tree.grid(row=0, column=0, sticky="nsew")
        y_scrollbar.grid(row=0, column=1, sticky="ns")
        
        # 配置tree_frame的网格权重
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)
        
        # 添加标签样式
        stats_tree.tag_configure('oddrow', background='#E8E8E8')
        stats_tree.tag_configure('evenrow', background='#F8F8F8')
        
        # 添加总计标签框架
        total_frame = ttk.Frame(main_frame)
        total_frame.pack(fill="x", padx=5, pady=5)
        
        # 添加总计标签
        total_label = ttk.Label(total_frame, 
                               text="总计 - 出库次数: 0 次    总出库数量: 0 件",
                               font=('Arial', 10, 'bold'))
        total_label.pack(side="right", padx=5)
        
        # 查询数据
        conn = sqlite3.connect('warehouse.db')
        cursor = conn.cursor()
        
        # 查询每个经办人的出库统计
        cursor.execute('''
        SELECT 
            t.operator,
            SUM(t.quantity) as total_quantity,
            COUNT(*) as transaction_count
        FROM transactions t
        WHERE t.product_id = ? 
            AND t.type = '出库'
            AND t.operator != ""
        GROUP BY t.operator
        ORDER BY total_quantity DESC
        ''', (product_id,))
        
        # 填充数据
        for i, row in enumerate(cursor.fetchall()):
            tag = 'evenrow' if i % 2 == 0 else 'oddrow'
            stats_tree.insert('', 'end', values=row, tags=(tag,))
        
        # 更新总计信息
        cursor.execute('''
        SELECT 
            SUM(quantity) as total_quantity,
            COUNT(*) as total_count
        FROM transactions
        WHERE product_id = ? 
            AND type = '出库'
            AND operator != ""
        ''', (product_id,))
        total_stats = cursor.fetchone()
        
        total_label.config(
            text=f"总计 - 出库次数: {total_stats[1] or 0} 次    总出库数量: {total_stats[0] or 0} 件"
        )
        
        conn.close()

    def on_tree_double_click(self, event):
        # 获取双击的项目和列
        item = self.trans_tree.selection()[0]
        column = self.trans_tree.identify_column(event.x)
        values = self.trans_tree.item(item)['values']
        
        # 根据双击的列显示不同的统计信息
        if column == '#2' or column == '#3':  # 商品ID或商品名称列
            self.show_product_stats(values[1], values[2])  # 传递商品ID和名称
        elif column == '#6':  # 经办人列
            self.show_operator_stats(values[5])  # 传递经办人姓名

    def show_product_stats(self, product_id, product_name):
        # 创建新窗口
        stats_window = tk.Toplevel(self.root)
        stats_window.title(f"商品 {product_name} 的出库统计")
        
        # 设置窗口大小和位置
        stats_window.geometry("600x400")
        x = self.root.winfo_x() + (self.root.winfo_width() - 600) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 400) // 2
        stats_window.geometry(f"+{x}+{y}")
        
        # 创建主框架
        main_frame = ttk.Frame(stats_window)
        main_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 创建日期选择框架
        date_frame = ttk.Frame(main_frame)
        date_frame.pack(fill="x", padx=5, pady=5)
        
        # 起始日期
        ttk.Label(date_frame, text="起始日期:").pack(side=tk.LEFT, padx=5)
        start_date = ttk.Entry(date_frame, width=12)
        start_date.pack(side=tk.LEFT, padx=5)
        start_date.insert(0, "2025-01-01")
        
        # 结束日期
        ttk.Label(date_frame, text="结束日期:").pack(side=tk.LEFT, padx=5)
        end_date = ttk.Entry(date_frame, width=12)
        end_date.pack(side=tk.LEFT, padx=5)
        end_date.insert(0, datetime.now().strftime('%Y-%m-%d'))
        
        # 创建表格框架
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 创建统计表格
        stats_tree = ttk.Treeview(tree_frame, 
                                 columns=("经办人", "总出库数量", "出库次数"),
                                 show="headings")
        
        # 设置表头
        stats_tree.heading("经办人", text="经办人")
        stats_tree.heading("总出库数量", text="总出库数量")
        stats_tree.heading("出库次数", text="出库次数")
        
        # 设置列宽
        stats_tree.column("经办人", width=100)
        stats_tree.column("总出库数量", width=80)
        stats_tree.column("出库次数", width=80)
        
        # 添加垂直滚动条
        y_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=stats_tree.yview)
        stats_tree.configure(yscrollcommand=y_scrollbar.set)
        
        # 布局表格和滚动条
        stats_tree.grid(row=0, column=0, sticky="nsew")
        y_scrollbar.grid(row=0, column=1, sticky="ns")
        
        # 配置tree_frame的网格权重
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)
        
        # 添加标签样式
        stats_tree.tag_configure('oddrow', background='#E8E8E8')
        stats_tree.tag_configure('evenrow', background='#F8F8F8')
        
        # 添加总计标签框架
        total_frame = ttk.Frame(main_frame)
        total_frame.pack(fill="x", padx=5, pady=5)
        
        # 添加总计标签
        total_label = ttk.Label(total_frame, 
                               text="总计 - 出库次数: 0 次    总出库数量: 0 件",
                               font=('Arial', 10, 'bold'))
        total_label.pack(side="right", padx=5)
        
        # 查询按钮
        def refresh_stats():
            try:
                # 验证日期格式
                datetime.strptime(start_date.get(), '%Y-%m-%d')
                datetime.strptime(end_date.get(), '%Y-%m-%d')
                
                # 清空现有数据
                for item in stats_tree.get_children():
                    stats_tree.delete(item)
                
                # 查询数据
                conn = sqlite3.connect('warehouse.db')
                cursor = conn.cursor()
                cursor.execute('''
                SELECT 
                    t.operator,
                    SUM(t.quantity) as total_quantity,
                    COUNT(*) as transaction_count
                FROM transactions t
                WHERE t.product_id = ? 
                    AND t.type = '出库'
                    AND t.date BETWEEN ? AND ?
                GROUP BY t.operator
                ORDER BY total_quantity DESC
                ''', (product_id, start_date.get(), end_date.get()))
                
                # 填充数据
                for i, row in enumerate(cursor.fetchall()):
                    tag = 'evenrow' if i % 2 == 0 else 'oddrow'
                    stats_tree.insert('', 'end', values=row, tags=(tag,))
                
                # 更新总计信息
                cursor.execute('''
                SELECT 
                    SUM(quantity) as total_quantity,
                    COUNT(*) as total_count
                FROM transactions
                WHERE product_id = ? 
                    AND type = '出库'
                    AND date BETWEEN ? AND ?
                ''', (product_id, start_date.get(), end_date.get()))
                total_stats = cursor.fetchone()
                
                total_label.config(
                    text=f"总计 - 出库次数: {total_stats[1] or 0} 次    总出库数量: {total_stats[0] or 0} 件"
                )
                
                conn.close()
            except ValueError:
                messagebox.showerror("错误", "日期格式错误，请使用YYYY-MM-DD格式")
        
        # 添加查询按钮
        ttk.Button(date_frame, text="查询", command=refresh_stats).pack(side=tk.LEFT, padx=20)
        
        # 初始加载数据
        refresh_stats()

    def show_operator_stats(self, operator):
        # 创建新窗口
        stats_window = tk.Toplevel(self.root)
        stats_window.title(f"经办人 {operator} 的出库统计")
        
        # 设置窗口大小和位置
        stats_window.geometry("800x600")  # 增加窗口尺寸
        x = self.root.winfo_x() + (self.root.winfo_width() - 800) // 2
        y = self.root.winfo_y() + (self.root.winfo_height() - 600) // 2
        stats_window.geometry(f"+{x}+{y}")
        
        # 创建主框架
        main_frame = ttk.Frame(stats_window)
        main_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 创建日期选择框架
        date_frame = ttk.Frame(main_frame)
        date_frame.pack(fill="x", padx=5, pady=5)
        
        # 起始日期
        ttk.Label(date_frame, text="起始日期:").pack(side=tk.LEFT, padx=5)
        start_date = DateEntry(date_frame, width=12, background='darkblue',
                              foreground='white', borderwidth=2,
                              date_pattern='yyyy-mm-dd',
                              year=2025)
        start_date.pack(side=tk.LEFT, padx=5)
        start_date.set_date("2025-01-01")  # 设置默认日期
        
        # 结束日期
        ttk.Label(date_frame, text="结束日期:").pack(side=tk.LEFT, padx=5)
        end_date = DateEntry(date_frame, width=12, background='darkblue',
                            foreground='white', borderwidth=2,
                            date_pattern='yyyy-mm-dd')
        end_date.pack(side=tk.LEFT, padx=5)
        end_date.set_date(datetime.now().strftime('%Y-%m-%d'))  # 设置为当前日期
        
        # 创建表格框架
        tree_frame = ttk.Frame(main_frame)
        tree_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        # 创建统计表格
        stats_tree = ttk.Treeview(tree_frame, 
                                 columns=("商品ID", "商品名称", "总出库数量", "出库次数"),
                                 show="headings")
        
        # 设置表头
        stats_tree.heading("商品ID", text="商品ID")
        stats_tree.heading("商品名称", text="商品名称")
        stats_tree.heading("总出库数量", text="总出库数量")
        stats_tree.heading("出库次数", text="出库次数")
        
        # 设置列宽
        stats_tree.column("商品ID", width=60)
        stats_tree.column("商品名称", width=200)
        stats_tree.column("总出库数量", width=80)
        stats_tree.column("出库次数", width=80)
        
        # 添加垂直滚动条
        y_scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=stats_tree.yview)
        stats_tree.configure(yscrollcommand=y_scrollbar.set)
        
        # 添加水平滚动条
        x_scrollbar = ttk.Scrollbar(tree_frame, orient="horizontal", command=stats_tree.xview)
        stats_tree.configure(xscrollcommand=x_scrollbar.set)
        
        # 布局表格和滚动条
        stats_tree.grid(row=0, column=0, sticky="nsew")
        y_scrollbar.grid(row=0, column=1, sticky="ns")
        x_scrollbar.grid(row=1, column=0, sticky="ew")
        
        # 配置tree_frame的网格权重
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)
        
        # 添加标签样式
        stats_tree.tag_configure('oddrow', background='#E8E8E8')
        stats_tree.tag_configure('evenrow', background='#F8F8F8')
        
        # 添加总计标签框架
        total_frame = ttk.Frame(main_frame)
        total_frame.pack(fill="x", padx=5, pady=5)
        
        # 添加总计标签
        total_label = ttk.Label(total_frame, 
                               text="总计 - 出库次数: 0 次    总出库数量: 0 件",
                               font=('Arial', 10, 'bold'))
        total_label.pack(side="right", padx=5)
        
        # 查询按钮
        def refresh_stats():
            try:
                # 获取日期字符串
                start_date_str = start_date.get_date().strftime('%Y-%m-%d')
                end_date_str = end_date.get_date().strftime('%Y-%m-%d')
                
                # 清空现有数据
                for item in stats_tree.get_children():
                    stats_tree.delete(item)
                
                # 查询数据
                conn = sqlite3.connect('warehouse.db')
                cursor = conn.cursor()
                cursor.execute('''
                SELECT 
                    t.product_id,
                    p.name,
                    SUM(t.quantity) as total_quantity,
                    COUNT(*) as transaction_count
                FROM transactions t
                LEFT JOIN products p ON t.product_id = p.id
                WHERE t.operator = ? 
                    AND t.type = '出库'
                    AND t.date BETWEEN ? AND ?
                GROUP BY t.product_id, p.name
                ORDER BY total_quantity DESC
                ''', (operator, start_date_str, end_date_str))
                
                # 填充数据
                for i, row in enumerate(cursor.fetchall()):
                    tag = 'evenrow' if i % 2 == 0 else 'oddrow'
                    stats_tree.insert('', 'end', values=row, tags=(tag,))
                
                # 更新总计信息
                cursor.execute('''
                SELECT 
                    SUM(quantity) as total_quantity,
                    COUNT(*) as total_count
                FROM transactions
                WHERE operator = ? 
                    AND type = '出库'
                    AND date BETWEEN ? AND ?
                ''', (operator, start_date_str, end_date_str))
                total_stats = cursor.fetchone()
                
                total_label.config(
                    text=f"总计 - 出库次数: {total_stats[1] or 0} 次    总出库数量: {total_stats[0] or 0} 件"
                )
                
                conn.close()
            except Exception as e:
                messagebox.showerror("错误", f"查询数据时出错：{str(e)}")
        
        # 添加查询按钮
        ttk.Button(date_frame, text="查询", command=refresh_stats).pack(side=tk.LEFT, padx=20)
        
        # 初始加载数据
        refresh_stats()

    def export_to_excel(self, table_type):
        try:
            # 选择保存位置
            file_path = filedialog.asksaveasfilename(
                defaultextension='.xlsx',
                filetypes=[("Excel files", "*.xlsx")],
                title="保存Excel文件"
            )
            
            if not file_path:
                return
            
            conn = sqlite3.connect('warehouse.db')
            
            if table_type == "products":
                # 导出商品列表
                df = pd.read_sql_query('''
                    SELECT id as 商品ID, name as 商品名称, 
                           quantity as 数量, price as 价格, 
                           category as 类别 
                    FROM products
                ''', conn)
            else:
                # 导出交易记录
                df = pd.read_sql_query('''
                    SELECT t.id as 交易ID, p.name as 商品名称,
                           t.type as 类型, t.quantity as 数量,
                           t.operator as 经办人, t.date as 日期
                    FROM transactions t
                    LEFT JOIN products p ON t.product_id = p.id
                    ORDER BY t.date DESC
                ''', conn)
            
            # 保存到Excel
            df.to_excel(file_path, index=False, sheet_name='数据')
            
            # 调整Excel格式
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment, Font
            
            wb = load_workbook(file_path)
            ws = wb.active
            
            # 设置列宽和格式
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
                
                # 设置标题行格式
                ws[f"{column_letter}1"].font = Font(bold=True)
                ws[f"{column_letter}1"].alignment = Alignment(horizontal="center")
            
            wb.save(file_path)
            conn.close()
            
            messagebox.showinfo("成功", "数据已成功导出到Excel文件！")
            
        except Exception as e:
            messagebox.showerror("错误", f"导出失败：{str(e)}")

    def on_transaction_click(self, event):
        # 获取当前选中的项目
        selection = self.trans_tree.selection()
        if not selection:
            return
            
        # 获取点击的列
        column = self.trans_tree.identify_column(event.x)
        values = self.trans_tree.item(selection[0])['values']
        
        # 如果点击的是商品ID或商品名称列
        if column in ('#2', '#3'):  # 商品ID或商品名称列
            # 填充商品ID和名称
            self.trans_product_id.delete(0, tk.END)
            self.trans_product_id.insert(0, values[1])  # 商品ID
            self.trans_product_name.delete(0, tk.END)
            self.trans_product_name.insert(0, values[2])  # 商品名称
        
        # 如果双击了经办人列
        elif column == '#6':  # 经办人列
            operator = values[5]
            if operator.strip():  # 只有当经办人不为空时才显示统计
                self.show_operator_stats(operator)

    def update_product_id_list(self, event=None):
        try:
            conn = sqlite3.connect('warehouse.db')
            cursor = conn.cursor()
            cursor.execute('SELECT id FROM products ORDER BY id')
            ids = [str(row[0]) for row in cursor.fetchall()]
            conn.close()
            
            # 更新下拉列表的值
            self.trans_product_id['values'] = ids
        except:
            pass

    def update_product_name_list(self, event=None):
        try:
            conn = sqlite3.connect('warehouse.db')
            cursor = conn.cursor()
            cursor.execute('SELECT name FROM products ORDER BY name')
            names = [row[0] for row in cursor.fetchall()]
            conn.close()
            
            # 更新下拉列表的值
            self.trans_product_name['values'] = names
        except:
            pass

    def update_product_name(self, event=None):
        try:
            product_id = self.trans_product_id.get()
            if product_id:
                conn = sqlite3.connect('warehouse.db')
                cursor = conn.cursor()
                cursor.execute('SELECT name FROM products WHERE id=?', (product_id,))
                result = cursor.fetchone()
                conn.close()
                
                # 清空名称输入框
                self.trans_product_name.delete(0, tk.END)
                
                # 只有在找到对应商品时才填充名称
                if result:
                    self.trans_product_name.insert(0, result[0])
                    
                    # 更新下拉列表
                    self.update_product_name_list()
        except:
            pass

    def update_product_id(self, event=None):
        try:
            product_name = self.trans_product_name.get()
            if product_name:
                conn = sqlite3.connect('warehouse.db')
                cursor = conn.cursor()
                cursor.execute('SELECT id FROM products WHERE name = ?', (product_name,))
                result = cursor.fetchone()
                conn.close()
                
                if result:
                    self.trans_product_id.delete(0, tk.END)
                    self.trans_product_id.insert(0, str(result[0]))
                    
                    # 更新下拉列表
                    self.update_product_id_list()
        except:
            pass

    def on_product_id_selected(self, event=None):
        try:
            product_id = self.trans_product_id.get()
            if product_id:
                conn = sqlite3.connect('warehouse.db')
                cursor = conn.cursor()
                cursor.execute('SELECT name FROM products WHERE id=?', (product_id,))
                result = cursor.fetchone()
                conn.close()
                
                if result:
                    self.trans_product_name.delete(0, tk.END)
                    self.trans_product_name.insert(0, result[0])
        except:
            pass

    def on_product_name_selected(self, event=None):
        try:
            product_name = self.trans_product_name.get()
            if product_name:
                conn = sqlite3.connect('warehouse.db')
                cursor = conn.cursor()
                cursor.execute('SELECT id FROM products WHERE name=?', (product_name,))
                result = cursor.fetchone()
                conn.close()
                
                if result:
                    self.trans_product_id.delete(0, tk.END)
                    self.trans_product_id.insert(0, str(result[0]))
        except:
            pass

    def clear_transaction_entries(self):
        """清空交易输入框"""
        self.trans_product_id.delete(0, tk.END)
        self.trans_product_name.delete(0, tk.END)
        self.trans_quantity.delete(0, tk.END)
        self.trans_operator.delete(0, tk.END)
        self.trans_date.delete(0, tk.END)
        self.trans_date.insert(0, datetime.now().strftime('%Y-%m-%d'))

    def show_about_info(self):
        """显示关于信息"""
        about_text = """
作者：奋青
邮箱：393283@qq.com

免责声明：
1. 本软件为免费开源软件，仅供学习和参考使用。

2. 使用者在使用本软件时需自行承担风险，作者不对使用本软件所导致的任何直接或间接损失负责。

3. 本软件不提供任何形式的保证，包括但不限于适销性、特定用途适用性的默示保证。

4. 使用者应自行负责数据的备份和安全，作者不对数据丢失或损坏承担责任。

5. 作者保留对本软件进行更新、修改或终止的权利，且无需事先通知。

6. 使用本软件即表示同意本免责声明的所有条款。

版权所有 © 2025 奋青
"""
        messagebox.showinfo("关于系统", about_text)

    def check_first_run(self):
        """检查是否首次运行程序"""
        config_file = 'app_config.txt'
        
        try:
            # 尝试读取配置文件
            with open(config_file, 'r') as f:
                # 文件存在，不是首次运行
                pass
        except FileNotFoundError:
            # 文件不存在，是首次运行
            # 显示免责声明
            self.show_about_info()
            
            # 创建配置文件标记已运行
            try:
                with open(config_file, 'w') as f:
                    f.write('initialized')
            except Exception as e:
                messagebox.showwarning("警告", f"无法创建配置文件：{str(e)}")

if __name__ == "__main__":
    try:
        root = tk.Tk()
        app = WarehouseSystem(root)
        root.mainloop()
    except Exception as e:
        print("程序启动时发生错误：")
        print("详细错误信息:", str(e))
        input("按回车键退出...")